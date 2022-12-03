from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import json

workbook = load_workbook(filename= '/home/mohammed/github/robotframework-soap-python/TestData/excels/DATASOURCE_INT.xlsx')
worksheet = workbook.active

my_list = []


last_column = len(list(worksheet.columns))
last_row    = len(list(worksheet.rows))


for row in range(1, last_row + 1):
    my_dict = {}
    for column in range(1, last_column + 1):
        column_letter = get_column_letter(column)
        if row > 1:
            my_dict[worksheet[column_letter + str(1)].value] = worksheet[column_letter + str(row)].value
    my_list.append(my_dict)

data = json.dumps(my_list, sort_keys=True, indent=4)
with open('/home/mohammed/github/robotframework-soap-python/TestData/json/DATASOURCE_INT.json', 'w') as f:
    f.write(data)